from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.chart import PieChart, PieChart3D, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl import load_workbook

class ExcelBot:
    def __init__(self, path, read_only=False, data_only=False):
        '''
        Parameters:
        - read_only (bool): True for gathering only values, not formulas
        '''
        self.wb = load_workbook(path, read_only=read_only, data_only=data_only)
        self.sheet = self.wb.active
        self.path=path
        self.sheetnames=self.wb.sheetnames

    def __enter__(self):
        self.wb = load_workbook(self.path, read_only=self.read_only)
        self.sheet = self.wb.active
        return self 
    
    def __exit__(self, exc_type, exc_value, traceback):
        if exc_type is not None:
            print(f"An exception occurred: {exc_value}")
        if self.wb:
            self.wb.close()
        return False

    def __len__(self):
        return self.sheet.max_row

    def __getitem__(self, position):
        row, column = position
        return self.sheet.cell(row=row, column=column).value

    def __str__(self):
        return f"Excel instance was created from the path {self.path}"

    def __contains__(self, item):
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value==item:
                    return True
        else: return False

    def change_sheet(self, sheet_name):
        """
        Change self.sheet

        Parameters:
        - sheet_name (str): Change self.sheet to sheet_name (e.g., "sheet1")
        """
        if sheet_name in self.wb.sheetnames:
            self.sheet = self.wb[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        
    def _get_sheet(self, sheet=None):
        """
        Stablish sheet to work with in functions, if sheet=None, use self.sheet
        """
        return sheet if sheet is not None else self.sheet
    
    def format_currency(self, col_letter, start_row=2, sheet=None):
        """
        Parameters:
        - col_letter (str): The letter of the column to format (e.g., "B").
        - start_row (int): The row from which formatting should begin. Default is 2.
        - sheet: The target sheet. Defaults to the active sheet.
        """
        sheet = self._get_sheet(sheet)
        col_index = self.get_column_index(col_letter)
        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '£#,##0.00'
    
    def create_chart(self, labels, datas, location="E30", title="Pie Chart", width=15, height=10, threeD=True, datalables=True, sheet=None):
        """
        Parameters:
        - labels (list): Labels location in the following format [min_col, min_row, max_row]
        - datas (list): Data location in the following format [min_col, min_row, max_row]
        - location (str): location of for the chart (e.g., "E30")
        - title (str): title of the chart
        - width (int): width of the chart
        - height (int): height of the chart
        - threeD (bool): True if 3D chart else 2D
        - datalables (bool): True if labels in the chart are desired
        """
        sheet=self._get_sheet(sheet)
        label=Reference(sheet, min_col=labels[0], min_row=labels[1], max_row=labels[2])
        data = Reference(sheet, min_col=datas[0], min_row=datas[1], max_row=datas[2]) 

        if threeD==True:
            self.pie = PieChart3D()
        else:  self.pie = PieChart()  
        self.pie.title = title
        self.pie.add_data(data, titles_from_data=False)  
        self.pie.set_categories(label)
        self.pie.layout = Layout(
            ManualLayout(
                x=0.04,
                y=0.1,
                w=0.65,
                h=0.65,
            )
        )

        self.pie.width = width  
        self.pie.height = height

        if datalables==True:
            dLbls = DataLabelList()
            dLbls.showVal = True
            dLbls.showCatName = False
            dLbls.showLeaderLines = True 
            self.pie.dLbls = dLbls

        sheet.add_chart(self.pie, location)

    def del_col(self, target, sheet=None):
        sheet=self._get_sheet(sheet)
        target_column_index = None
        for col_index, cell in enumerate(sheet[1], 1):
            if cell.value == target:
                target_column_index = col_index
                break  
        if target_column_index:
            sheet.delete_cols(target_column_index)

    def get_indexes_from_target(self, parameter, target_col, sheet=None):
        indexes = []
        sheet = self._get_sheet(sheet)
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), start=1):
            cell_value = row[target_col - 1].value
            if cell_value is not None:
                if isinstance(parameter, list):
                    if any(section.lower() in str(cell_value).lower() for section in parameter):
                        indexes.append(row_index)
                elif isinstance(parameter, str):
                    if parameter.strip().lower() == str(cell_value).strip().lower():
                        indexes.append(row_index)
                else:
                    if parameter == cell_value:
                        indexes.append(row_index)
        return indexes

    def add_cols(self, columns, row=1, sheet=None):
        sheet=self._get_sheet(sheet)
        max_col = sheet.max_column
        if isinstance(columns, list):
            for column_index, string in enumerate(columns):
                insertion_index = max_col + column_index + 1
                sheet.insert_cols(insertion_index, 1)
                sheet.cell(row=row, column=insertion_index, value=string)
        else:
            sheet.insert_cols(max_col + 1, 1)
            sheet.cell(row=row, column=max_col + 1, value=columns)

    def get_column_index(self, col_str):
        index = 0
        for char in col_str:
            index = index * 26 + (ord(char.upper()) - ord('A') + 1)
        return index

    def copy_paste(self, source_range, target_cell, source_sheet=None, target_sheet=None, vals_only=False):
        '''
        Parameters:
        - source_range (str): The range of cells to copy, in Excel-style notation (e.g., "A1:C3" for a range from cell A1 to C3).
        - target_cell (str): The top-left cell in the target sheet where the copied data will be pasted, in Excel-style notation (e.g., "D1").
        '''
        if source_sheet==None:
            source_sheet=self.sheet
        if target_sheet==None:
            target_sheet=self.sheet
        if not isinstance(target_sheet, type(self.wb.active)):
            print("Target sheet is not a valid Worksheet. Please check the sheet reference.")
            return

        if isinstance(source_range, str):
            items=source_range.split(":")
            start=items[0]
            end=items[1]
            start_col_str = ''.join([ch for ch in start if ch.isalpha()])
            start_row_str = ''.join([ch for ch in start if ch.isdigit()]) 
            start_col=self.get_column_index(start_col_str)
            end_col_str = ''.join([ch for ch in end if ch.isalpha()])
            end_row_str = ''.join([ch for ch in end if ch.isdigit()])
            end_col=self.get_column_index(end_col_str)
            values=[]
            for row in source_sheet.iter_rows(min_row=int(start_row_str), max_row=int(end_row_str), min_col=int(start_col), max_col=int(end_col), values_only=vals_only):
                if not vals_only:
                    values.append([cell.value for cell in row])
                else:
                    values.append(row)

        else: print("source_range must be a string")

        if isinstance(target_cell, str):
            target_str_col="".join(char for char in target_cell if char.isalpha())
            target_str_row="".join(char for char in target_cell if char.isdigit())
            target_col=self.get_column_index(target_str_col)
            for i, row_values in enumerate(values):
                for j, value in enumerate(row_values):
                    target_sheet.cell(row=int(target_str_row) + i, column=target_col + j, value=value)

        else: print("target_cell must be a string")

    def replace_data(self, values, target_cell, target_sheet=None):
        '''
        Parameters:
        - values (list): List of list (eg. [['Name', 'Total Money', 'Total Hours'], ['Back Corridor', None, None], ['Mccann, Thomas', 0, 8.25]])
        - target_cell (str): The top-left cell in the target sheet where the copied data will be pasted, in Excel-style notation (e.g., "D1").
        '''
        if target_sheet==None:
            target_sheet=self.sheet
        if isinstance(values, list):
            pass
        else: print("Values must be a list of list")
        if isinstance(target_cell, str):
            target_str_col="".join(char for char in target_cell if char.isalpha())
            target_str_row="".join(char for char in target_cell if char.isdigit())
            target_col=self.get_column_index(target_str_col)
            for i, row_values in enumerate(values):
                for j, value in enumerate(row_values):
                    target_sheet.cell(row=int(target_str_row) + i, column=target_col + j, value=value)
        else: print("target_cell must be a string")

    def get_column_letter(self, col_index):
        letter = ''
        while col_index > 0:
            col_index, remainder = divmod(col_index - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter

    def auto_adjust_width(self, sheet=None):
        sheet=self._get_sheet(sheet)
        for i in range(1, sheet.max_column+1):
            col_letter = self.get_column_letter(i)
            longest_val=0
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=i, max_col=i):
                for cell in row:
                    if cell.value:
                        length = len(str(cell.value)) if isinstance(cell.value, (int, float)) else len(str(cell.value.strip()))
                        if length > longest_val:
                            longest_val = length
                            addition = 0.85 if isinstance(cell.value, int) else 1.1
            col_length = longest_val * addition
            sheet.column_dimensions[col_letter].width = col_length

    def add_sum_cell(self, col_letter, start=1, end=None, sheet=None):
        '''
        Parameters:
        - col_letter (str): (eg. "A")
        '''
        if isinstance(col_letter, str):
            col_idx=self.get_column_index(col_letter)
        else: print("col_letter must be a string")
        sheet=self._get_sheet(sheet)
        if end is None:
            end = sheet.max_row

        formula=f"=SUM({col_letter}{start}:{col_letter}{end})"
        sheet.cell(row=end, column=col_idx, value=formula)

    def add_average_cell(self, target_col, start=1, end=None, sheet=None):
        '''
        Parameters:
        - target_col (str): (eg. "A")
        '''
        sheet=self._get_sheet(sheet)
        if end is None:
            end = sheet.max_row
        col_idx=self.get_column_index(target_col)
        formula=f"=AVERAGE({target_col}{start}:{target_col}{end-1})"
        sheet.cell(row=end, column=col_idx, value=formula)
        
    def get_border_formatting(self, target_cell, sheet=None):
        """
        Parameters:
        - target_cell (str): Cell address (e.g., "A1").
        """
        sheet=self._get_sheet(sheet)
        cell=sheet[target_cell]

        border_styles = {
        "left": cell.border.left.style if cell.border.left else "None",
        "right": cell.border.right.style if cell.border.right else "None",
        "top": cell.border.top.style if cell.border.top else "None",
        "bottom": cell.border.bottom.style if cell.border.bottom else "None"
        }

        return border_styles
    
    def get_column_data(self, col_name):
        """Returns all data in a specified column."""
        for cell in self.sheet[1]:
            if cell.value == col_name:
                col_index = cell.column
                data = [self.sheet.cell(row=row, column=col_index).value for row in range(2, self.sheet.max_row + 1)]
                print(f"Data in '{col_name}' column:", data)
                return data
        print(f"Column '{col_name}' not found.")
    
    def copy_paste_border_formatting(self, source_cell, target, sheet=None):
        """
        Parameters:
        - source_cell (str): Cell to copy style from (e.g., "A1"). 
        - target (str): Cell of number of cells where the formatting wants to be added. (e.g., "A1" or for range "A1:C3").
        """
        sheet=self._get_sheet(sheet)

        border_styles=self.get_border_formatting(source_cell)
        border = Border(
        left=Side(style=border_styles["left"]) if border_styles["left"] != "None" else Side(),
        right=Side(style=border_styles["right"]) if border_styles["right"] != "None" else Side(),
        top=Side(style=border_styles["top"]) if border_styles["top"] != "None" else Side(),
        bottom=Side(style=border_styles["bottom"]) if border_styles["bottom"] != "None" else Side()
        )

        if ":" in target:
            cells = sheet[target]
        else: 
            cells = [[sheet[target]]]
        
        for row in cells:
            for cell in row:
                cell.border = border

    def apply_alignment(self, direction, target, sheet=None):
        '''
        Parameters:
        - direction (str): "left", "center", "right"
        - target (str): Desired cells for alignment (e.g., "A1" for range "A1:C3" for all "all")
        '''
        sheet=self._get_sheet(sheet)

        alignments = {
        "left": Alignment(horizontal="left"),
        "center": Alignment(horizontal="center"),
        "right": Alignment(horizontal="right")
        }
        if direction not in alignments:
            raise ValueError("Invalid alignment direction. Choose 'left', 'center', or 'right'.")

        alignment = alignments[direction]
        if ":" in target:
            cells = sheet[target]
        elif target.lower() == "all":
            cells = sheet.iter_rows()
        else:
            cells = [[sheet[target]]]
        
        for row in cells:
            for cell in row:
                cell.alignment = alignment

    def apply_font_formatting(self, target, font, font_size, bold=False, italic=False, color="000000", sheet=None):
        '''
        Parameters:
        - target (str): Desired cells for alignment (e.g., "A1" for range "A1:C3" for all "all")
        - font (str): Desired font style (e.g., "Calibri")
        - font_size (int): Desired letter size (e.g., "11")
        - bold (bool): True for bold text
        - italic (bool): True for italic text
        - color (str): Desired color (e.g., "000000")
        '''
        sheet=self._get_sheet(sheet)

        font=Font(name=font, size=font_size, bold=bold, italic=italic, color=color)
        if ":" in target:
            cells=sheet[target]
        elif target=="all":
            cells=sheet.iter_rows()
        else:
            cells=[[sheet[target]]]

        for rows in cells:
            for cell in rows:
                cell.font=font
        
    def clear_cell_formatting(self, target, sheet=None):
        ''' 
        Parameters:
        - target (str): Target cell or range (e.g., "A1" or "A1:C3").
        '''
        sheet=self._get_sheet(sheet)
        if ":" in target:
            cells = sheet[target]
        elif target=="all":
            cells=sheet.iter_rows()
        else:
            cells = [[sheet[target]]]

        for row in cells:
            for cell in row:
                cell.font = Font()
                cell.fill = PatternFill()
                cell.alignment = Alignment()
                cell.border = Border()

    def find_max(self, target, sheet=None):
        ''' 
        Parameters:
        - target (str): Target col (e.g., "A").
        '''
        sheet=self._get_sheet(sheet)
        col_idx=self.get_column_index(target)
        max_val=0
        index=-1
        for idx, row in enumerate(sheet.iter_rows(0, sheet.max_row, col_idx, col_idx)):
            cell = row[0]
            if isinstance(cell.value, (int, float)):
                    if cell.value>max_val:
                        max_val=cell.value
                        index=idx
            else: continue
        if index==-1:
            raise TypeError(f"No integer values found in column '{target}' in the specified sheet.")
        return (index, max_val)

    def find_min(self, target, sheet=None):
        ''' 
        Parameters:
        - target (str): Target col (e.g., "A").
        '''
        sheet=self._get_sheet(sheet)
        
        col_idx=self.get_column_index(target)
        min_val=float('inf')
        index=-1
        for idx, row in enumerate(sheet.iter_rows(0, sheet.max_row, col_idx, col_idx)):
            cell = row[0]
            if isinstance(cell.value, (int, float)):
                    if cell.value<min_val:
                        min_val=cell.value
                        index=idx
            else: continue
        if index==-1:
            raise TypeError(f"No integer values found in column '{target}' in the specified sheet.")
        return (index, min_val)
    
    def count_vals(self, target_val, target_col, sheet=None):
        ''' 
        Parameters:
        - target_val (str): Value being search for (e.g., 0)
        - target_col (str): Target col (e.g., "A").
        '''
        sheet=self._get_sheet(sheet)
        count=0
        col_idx=self.get_column_index(target_col)
        if sheet is None:
            sheet = self.sheet
        for row in sheet.iter_rows(0, sheet.max_row, col_idx, col_idx):
            cell = row[0]
            if isinstance(cell.value, type(target_val)):
                    if cell.value==target_val:
                        count+=1
        return count
    
    def filter_val(self, filter, source, target_col, sheet=None):
        ''' 
        Parameters:
        - source (int): Reference val (e.g., 3)
        - filter (str): Filter (e.g., target "<" source)
        - target_col (str): Target col (e.g., "A").
        '''
        sheet=self._get_sheet(sheet)
        filered_vals=[]
        col_idx=self.get_column_index(target_col)
        if sheet is None:
            sheet = self.sheet
        for idx, row in enumerate(sheet.iter_rows(0, sheet.max_row, col_idx, col_idx)):
            cell = row[0]
            if isinstance(cell.value, (int, float)):
                if filter=="<":
                    if cell.value<source:
                        filered_vals.append((idx, cell.value))
                elif filter=="<=":
                    if cell.value<=source:
                        filered_vals.append((idx, cell.value))
                elif filter==">":
                    if cell.value>source:
                        filered_vals.append((idx, cell.value))
                elif filter==">=":
                    if cell.value>=source:
                        filered_vals.append((idx, cell.value))
                elif filter=="=":
                    if cell.value==source:
                        filered_vals.append((idx, cell.value))

        return filered_vals
    
    def math_calc(self, operation, source, target_col, destination, sheet=None):
        '''
        Parameters:
        - operation (str): Mathematical operation to be used (e.g., "+", "-", "x", "/")
        - source (str, int, float): The number, column or column range to operate with (e.g., "A", "A1:A5", 6)
        - target_col (str): The column or column range to apply the operation (e.g., "A", "A1:A5")
        - destination (str): The cell to paste the mathematical procedures
        '''
        sheet=self._get_sheet(sheet)

        operations = {
            '+': '+',
            '-': '-',
            'x': '*',
            '/': '/'
        }
        operation_symbol = operations.get(operation)
        if operation_symbol is None:
            raise ValueError(f"Invalid operation '{operation}' provided.")

        if isinstance(source, (int, float)):
            source_values = [source] * sheet.max_row
        elif isinstance(source, str):
            if ":" in source:
                source_values = [cell.coordinate for row in sheet[source] for cell in row]
            else:
                source_values = [sheet[source].coordinate]

        if ":" in target_col:
            target_cells = [cell.coordinate for row in sheet[target_col] for cell in row]
        else:
            target_cells = [sheet[target_col].coordinate]

        dest_col = self.get_column_index(destination[0])
        dest_row = int(destination[1:])

        for i, (src_cell, tgt_cell) in enumerate(zip(source_values, target_cells)):
            if isinstance(src_cell, (int, float)):
                formula = f'={tgt_cell}{operation_symbol}{src_cell}'
            else:
                formula = f'={src_cell}{operation_symbol}{tgt_cell}'
            sheet.cell(row=dest_row + i, column=dest_col, value=formula)

    def remove_duplicates(self, mode=0, column_letter=None, delete=False, tolerance=0, sheet=None):
        '''
        Parameters:
        - mode (int): The mode is used for applying the remove dupicates to a designated area. 0 for columns, 1 for row and columns and 2 for identical rows
        - column_letter (str): The letter of the column (e.g., "A")
        - tolerance (int): The number of duplicates allowed before removal
        '''
        values=[]
        sheet=self._get_sheet(sheet)
        if mode==0:
            if column_letter==None:
                raise ValueError("column_letter can not be none for mode 0")
            column_idx=self.get_column_index(column_letter)
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=column_idx, max_col=column_idx):
                cell=row[0]
                count = values.count(cell.value)
                values.append(cell.value)
                if count>tolerance:
                    cell.value=""
        elif mode==1:
            for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_column):
                for cell in row:
                    count = values.count(cell.value)
                    values.append(cell.value)
                    if count>tolerance:
                        cell.value=""
        elif mode == 2:
            row_vals = []
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                row_vals.append([cell.value for cell in row])

            row_count = {}
            for row_data in row_vals:
                row_tuple = tuple(row_data)
                if row_tuple in row_count:
                    row_count[row_tuple] += 1
                else:
                    row_count[row_tuple] = 1

            rows_to_delete = []
            for idx, row_data in enumerate(row_vals):
                if row_count[tuple(row_data)] > tolerance and row_count[tuple(row_data)] > 1:
                    rows_to_delete.append(idx + 1)
                    row_count[tuple(row_data)] -= 1
            print(rows_to_delete)

            if delete==True:
                for row_idx in reversed(rows_to_delete):
                    sheet.delete_rows(row_idx)
            else:
                for row_idx in reversed(rows_to_delete):
                    for row in sheet.iter_rows(min_row=row_idx, max_row=row_idx, min_col=1, max_col=sheet.max_column):
                        for cell in row:
                            cell.value=""